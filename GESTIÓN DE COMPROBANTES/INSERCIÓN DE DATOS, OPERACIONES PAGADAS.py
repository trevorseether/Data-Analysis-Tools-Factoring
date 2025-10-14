# -*- coding: utf-8 -*-
"""
Created on Mon Apr  7 10:42:50 2025

@author: Joseph Montoya
"""

# pip install boto3 pyathena 

import pandas as pd
# import numpy as np
# import boto3
from pyathena import connect
# import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import os

import shutil
from datetime import datetime

import warnings
warnings.filterwarnings("ignore")

#%% lectura de archivo de garantías negativas:
garantias = pd.read_excel(r'C:/Users/Joseph Montoya/Downloads/11.10.25 - CUADRE DE CAJA.xlsx',
                          sheet_name = 'Desembolso de Garantías')
garantias = garantias[ ~pd.isna(garantias ['CODIGO SUBASTA']) ]
garantias = garantias[ ~pd.isna(garantias ['GARANTIA']) ]
garantias['GARANTIA'] = garantias['GARANTIA'].astype(str)
garantias = garantias[ garantias['GARANTIA'] != "#VALUE!" ]
garantias = garantias[['CODIGO SUBASTA', 'GARANTIA']]

garantias['GARANTIA'] = pd.to_numeric(garantias['GARANTIA'])
garantias = garantias[ garantias['GARANTIA'] < 0 ]

#%% Credenciales de AmazonAthena
import json
with open(r"C:/Users/Joseph Montoya/Desktop/credenciales actualizado.txt") as f:
    creds = json.load(f)

conn = connect(
    aws_access_key_id     = creds["AccessKeyId"],
    aws_secret_access_key = creds["SecretAccessKey"],
    aws_session_token     = creds["SessionToken"],
    s3_staging_dir        = creds["s3_staging_dir"],
    region_name           = creds["region_name"]
    
    )

#%%
# Ejecutar consulta (fac_ingresos_teoricos_contabilidad_vesquivel)

query = """

with Pagos as (
	select b.code as Subasta,
		--Codigo subasta
		max(c.date) as Fecha_Pago,
		--Fecha de Pago máximo, se debe restar 5 horas
		sum(c.amount) as Monto_Pagado,
		--Monto Pagado por el cliente
		sum(c.distribution_provider_amount_payment_client) AS monto_pagado_total
	from prod_datalake_analytics.fac_client_payments a
		left join prod_datalake_analytics.fac_auctions b on a.auction_id = b._id
		left join prod_datalake_analytics.fac_client_payment_payments c on a._id = c.client_payment_id
	where a.status = 'finalizado'
	group by b.code
),
Pagos_crowd as (
	select auction_code as Subasta,
		ROUND(sum(interes_payment), 2) as Monto_pagado_crowd
	from prod_datalake_analytics.fac_client_payment_investors
	where bid_created_at >= TIMESTAMP '2024-11-01 00:00:00.000'
	group by auction_code
),
payments1 as (
	SELECT *,
		b.code as codigo,
		(
			CASE
				WHEN (
					COALESCE(fcpp.distribution_provider_amount_bussinesman, 0) > 0
				) THEN fcpp.distribution_provider_amount_bussinesman
				WHEN (
					(
						COALESCE(fcpp.pay_order_businessman_amount, 0) > 0
					)
					AND (COALESCE(fcpp.distribution_provider_igv, 0) > 0)
				) THEN 0
				WHEN (
					(
						COALESCE(fcpp.pay_order_businessman_amount, 0) > 0
					)
					AND (COALESCE(fcpp.distribution_provider_igv, 0) = 0)
				) THEN pay_order_businessman_amount ELSE 0
			END
		) guarantee_paid
	FROM "prod_datalake_analytics"."fac_client_payment_payments" as fcpp
		left join prod_datalake_analytics.fac_client_payments as c on fcpp.client_payment_id = c._id
		left join prod_datalake_analytics.fac_requests as b on c.request_id = b._id
	where b.status = 'closed'
		and c.status = 'finalizado' --order by b.code
),
int_pagado as(
	select a.codigo,
		round(
			sum(
				(
					CASE
						WHEN (
							(COALESCE(pay_order_businessman_amount, 0) > 0)
							AND (COALESCE(distribution_provider_igv, 0) > 0)
						) THEN (
							COALESCE(a.distribution_provider_INTeres_payment, 0) + COALESCE(distribution_provider_igv, 0)
						)
						WHEN (
							(COALESCE(pay_order_businessman_amount, 0) = 0)
							AND (COALESCE(distribution_provider_igv, 0) > 0)
						) THEN (
							COALESCE(a.distribution_provider_INTeres_payment, 0) + COALESCE(distribution_provider_igv, 0)
						)
						WHEN (
							(
								COALESCE(a.distribution_provider_INTeres_payment, 0) = 0
							)
							AND (
								COALESCE(
									a.distribution_provider_amount_capital_payment,
									0
								) = 0
							)
							AND COALESCE((guarantee_paid = 0))
						) THEN a.distribution_provider_INTeres ELSE COALESCE(a.distribution_provider_INTeres_payment, 0)
					END
				)
			),
			2
		) interest_paid,
		
		sum( coalesce (diff_interest_proforma_interes_total_real,0))  as int_pronto_pago
		
	from payments1 as a
	group by a.codigo
),
estado_cob as (
	select fr.code,
		count(fr.code) as conteo_inversionistas,
		count(aux.cobrados) as conteo_cobrados,
		case
			when count(fr.code) = count(aux.cobrados) then 'Cerrado' else ''
		end as "Estado de cobranza admin"
	from prod_datalake_analytics.fac_bids as fb
		left join prod_datalake_analytics.fac_requests as fr on fr._id = fb.request_id
		left join (
			select fr.code,
				count(fb.pay_status) as cobrados
			from prod_datalake_analytics.fac_bids as fb
				left join prod_datalake_analytics.fac_requests as fr on fr._id = fb.request_id
			where fb.status = 'ganado'
				and fb.pay_status = 'cobrado'
			group by fr.code
		) as aux on aux.code = fr.code
	where fb.status = 'ganado'
	group by fr.code
)
select case
		when proforma_strategy_name = 'factoring-v1-new' then 'Estrategias' else ''
	end as Flag_Estrategias,
	a.code as Subasta,
	--subasta
	a.status as Status,
	proforma_simulation_currency as Moneda,
	--moneda
	DATE(interest_proforma_disbursement_date) as Fecha_desembolso,
	--fecha desembolso registrado admin
	DATE(
		date_add('hour', -5, CAST(a.closed_at AS timestamp))
	) as Fecha_venta,
	--fecha de venta
	DATE(proforma_client_payment_date_expected) as Fecha_esperada_pago,
	--Fecha de pago proveedor
	proforma_simulation_financing_total as Monto_Financiado,
	--Monto Financiado
	proforma_simulation_net as Monto_neto,
	--Monto Neto Factura
	proforma_financing_interest_rate as Tasa_interes_empresario,
	--tasa de interés
	proforma_profit_interest_rate as Tasa_interes_crowd,
	--tasa de interés crowd
	interest_proforma_simulation_financing_cost_value as Costo_Financiamiento,
	--proforma_simulation_financing_commission as comision_sin_igv, --Comision sin IGV
	--proforma_simulation_financing_commission_igv as igv_comision, --IGV comisión
	--proforma_simulation_financing_commission + proforma_simulation_financing_commission_igv as comision_total, --comisión con IGV
	a.company_ruc as Ruc_proveedor,
	a.company_name as Razon_Social,
	g.address as Direccion,
	a.businessman_email as Correo,
	b.code as Comprobante_Comision,
	c.code as Comprobante_interes,
	DATE(d.Fecha_desembolso) as Fecha_Desembolso_Hubspot,
	DATE(e.Fecha_Pago) as Fecha_Pago_real,
	--e.Monto_Pagado as Monto_pagado
	e.monto_pagado_total as "Monto pagado total",
	f.Monto_pagado_crowd as Monto_pagado_crowd,
	ec."Estado de cobranza admin",
	ip.interest_paid as "Costo de Financiamiento cobrado admin",
	ip.int_pronto_pago,
	'grupo 1' as origen_query
from prod_datalake_analytics.fac_requests a
	left join (
		select *
		from prod_datalake_analytics.view_prestamype_fac_cpe
		where concept = 'commission-factoring'
	) b on a._id = b.request_id
	left join (
		(
			select vpcp.*
			from prod_datalake_analytics.view_prestamype_fac_cpe as vpcp
				inner join (
					select request_id,
						min(created_at) as created_at
					from prod_datalake_analytics.view_prestamype_fac_cpe
					where concept = 'interest-factoring'
					group by request_id
				) as vpcpmax on vpcp.request_id = vpcpmax.request_id
				and vpcp.created_at = vpcpmax.created_at
			where vpcp.concept = 'interest-factoring'
		)
	) c on a._id = c.request_id
	left join (
		select c_digo_de_subasta,
			min(fecha_de_desembolso_registrado) as Fecha_desembolso
		from prod_datalake_master.hubspot__pagos_facturas
		where hs_pipeline = '42653778'
        
		group by c_digo_de_subasta
	) d on a.code = d.c_digo_de_subasta
	left join Pagos as e on a.code = e.Subasta
	left join Pagos_crowd as f on a.code = f.Subasta
	left join prod_datalake_analytics.view_fac_companies as g on a.company_ruc = g.ruc
	left join int_pagado as ip on ip.codigo = a.code
	left join estado_cob as ec on ec.code = e.Subasta
where product = 'factoring'
	and a.status in('closed', 'ongoing')
	and interest_proforma_disbursement_date is null --and DATE(e.Fecha_Pago) is not null
	and proforma_strategy_name = 'factoring-v1-new'
union all
select case
		when proforma_strategy_name = 'factoring-v1-new' then 'Estrategias' else ''
	end as Flag_Estrategias,
	a.code as Subasta,
	--subasta
	a.status as Status,
	currency as Moneda,
	--moneda
	DATE(interest_proforma_disbursement_date) as Fecha_desembolso,
	--fecha desembolso registrado admin
	DATE(
		date_add('hour', -5, CAST(a.closed_at AS timestamp))
	) as Fecha_venta,
	--fecha de venta
	DATE(interest_proforma_client_payment_date_expected) as Fecha_esperada_pago,
	--Fecha de pago proveedor
	interest_proforma_simulation_financing_total as Monto_Financiado,
	--Monto Financiado
	interest_proforma_simulation_net as Monto_neto,
	--Monto Neto Factura
	interest_proforma_financing_interest_rate as Tasa_interes_empresario,
	--tasa de interés
	proforma_profit_interest_rate as Tasa_interes_crowd,
	--tasa de interés crowd
	interest_proforma_simulation_financing_cost_value as Costo_Financiamiento,
	--interest_proforma_simulation_financing_commission as comision_sin_igv, --Comision sin IGV
	--interest_proforma_simulation_financing_commission_igv as igv_comision, --IGV comisión
	--interest_proforma_simulation_financing_commission + interest_proforma_simulation_financing_commission_igv as comision_total, --comisión con IGV
	a.company_ruc as Ruc_proveedor,
	a.company_name as Razon_Social,
	g.address as Direccion,
	a.businessman_email as Correo,
	b.code as Comprobante_Comision,
	c.code as Comprobante_interes,
	DATE(d.Fecha_desembolso) as Fecha_Desembolso_Hubspot,
	DATE(e.Fecha_Pago) as Fecha_Pago_real,
	e.monto_pagado_total as "Monto pagado total",
	f.Monto_pagado_crowd as Monto_pagado_crowd,
	ec."Estado de cobranza admin",
	ip.interest_paid as "Costo de Financiamiento cobrado admin",
	ip.int_pronto_pago,
	--e.Monto_Pagado as Monto_pagado
	'grupo 2' as origen_query
from prod_datalake_analytics.fac_requests a
	left join (
		select *
		from prod_datalake_analytics.view_prestamype_fac_cpe
		where concept = 'commission-factoring'
	) b on a._id = b.request_id
	left join (
		(
			select vpcp.*
			from prod_datalake_analytics.view_prestamype_fac_cpe as vpcp
				inner join (
					select request_id,
						min(created_at) as created_at
					from prod_datalake_analytics.view_prestamype_fac_cpe
					where concept = 'interest-factoring'
					group by request_id
				) as vpcpmax on vpcp.request_id = vpcpmax.request_id
				and vpcp.created_at = vpcpmax.created_at
			where vpcp.concept = 'interest-factoring'
		)
	) c on a._id = c.request_id
	left join (
		select c_digo_de_subasta,
			min(fecha_de_desembolso_registrado) as Fecha_desembolso
		from prod_datalake_master.hubspot__pagos_facturas
		where hs_pipeline = '42653778'
		group by c_digo_de_subasta
	) d on a.code = d.c_digo_de_subasta
	left join Pagos as e on a.code = e.Subasta
	left join Pagos_crowd as f on a.code = f.Subasta
	left join prod_datalake_analytics.view_fac_companies as g on a.company_ruc = g.ruc
	left join int_pagado as ip on ip.codigo = a.code
	left join estado_cob as ec on ec.code = e.Subasta
where product = 'factoring'
	and a.status in('closed', 'ongoing')
	and interest_proforma_disbursement_date is not null --and DATE(e.Fecha_Pago) is not null
	and proforma_strategy_name = 'factoring-v1-new'



"""

cursor = conn.cursor()
cursor.execute(query)

# Obtener los resultados
resultados = cursor.fetchall()

# Obtener los nombres de las columnas
column_names = [desc[0] for desc in cursor.description]

# Convertir los resultados a un DataFrame de pandas
df = pd.DataFrame(resultados, columns=column_names)

del df['origen_query']

print('datos obtenidos de Query principal')

duplicados = df[df["Subasta"].duplicated()]
if duplicados.shape[0] > 0 :
    print('alerta de duplicados')


#%%
#Cambiar todas las columnas de fecha a formato fecha:
for col in df.filter(like="Fecha").columns:
    df[col] = pd.to_datetime(df[col], errors='coerce')

#Ordernar por Fecha de Venta
df = df.sort_values(by='Fecha_venta', ascending = True).reset_index(drop = True)

#Cambiar todas las columnas en texto a formato numérico:
columnas_a_convertir =['Monto_Financiado',
                       'Monto_neto',
                       'Tasa_interes_empresario',
                       'Tasa_interes_crowd',
                       'Costo_Financiamiento']

df[columnas_a_convertir] = df[columnas_a_convertir].apply(pd.to_numeric)

#%%
UBI       = 'G:/Mi unidad'
DOCUMENTO = 'Pagados 122024 en adelante.xlsx' #'Pagados arreglado.xlsx'
compartido=pd.read_excel(UBI + '/' + DOCUMENTO,
                         sheet_name ='Online')

nuevos_registros = df[~df['Subasta'].isin(compartido['Subasta'])]

#%% creación de copia de seguridad
hoy_formateado = datetime.today().strftime('%d-%m-%Y')  # o '%Y-%m-%d', etc.

# Ruta base del nuevo archivo
nombre_base = f'Pagados {hoy_formateado}.xlsx'
ruta_base = os.path.join(UBI, 'backups PAGADOS')
destination_path = os.path.join(ruta_base, nombre_base)

# Verificar si el archivo ya existe y generar uno con (1), (2), etc.
contador = 1
while os.path.exists(destination_path):
    nombre_con_sufijo = f'Pagados {hoy_formateado} ({contador}).xlsx'
    destination_path = os.path.join(ruta_base, nombre_con_sufijo)
    contador += 1

# Copiar el archivo
source_path = os.path.join(UBI, DOCUMENTO)
shutil.copy(source_path, destination_path)

#%%
compartido['Subasta'] = compartido['Subasta'].str.strip()
df['Subasta'] = df['Subasta'].str.strip()

df_para_merge = df.copy()
del df_para_merge['Flag_Estrategias']

base_actual  = compartido[['Flag_Estrategias', 'Subasta']].merge(df_para_merge,
                                                                 on  = 'Subasta',
                                                                 how = 'left')
del base_actual['Flag_Estrategias']

#%% VALIDACIÓN DE DUPLICADOS
base_actual = base_actual[~((base_actual['Subasta'] == 'KZaM2xli') & (base_actual['Comprobante_interes'] == 'F002-26532'))]

dupli = base_actual[base_actual.duplicated(subset=['Subasta'])]
if dupli.shape[0] > 0: 
    print('alerta de duplicados, urgente validar')
    print(dupli)

#%% incluir columnas que deben ser obtenidas del archivo original
base_actual = base_actual.merge(compartido[['Subasta',
                                            'Estado de cobranza real']],
                                on  = 'Subasta',
                                how ='left')

copia_base = base_actual.copy()
#%%
# elimina la primera columna del dataframe 
# base_actual  = base_actual.iloc[:, range(1, base_actual.shape[1])]

base_actual  = base_actual[['Status',
                            'Moneda', 
                            'Fecha_desembolso', 
                            'Fecha_venta',     
                            'Fecha_esperada_pago', 
                            'Monto_Financiado', 
                            'Monto_neto',   
                            'Tasa_interes_empresario', 
                            'Tasa_interes_crowd',
                            'Costo_Financiamiento',
                            'Ruc_proveedor',
                            'Razon_Social',
                            'Direccion',
                            'Correo',
                            'Comprobante_Comision',
                            'Comprobante_interes', 
                            'Fecha_Desembolso_Hubspot', 
                            'Fecha_Pago_real',
                            'Monto pagado total', # nuevo
                            'Estado de cobranza real', # nuevo, llenado por tesorería
                            'Estado de cobranza admin', # nuevo, implementado
                            'Monto_pagado_crowd', # Interés Bruto pagado a Crowd (Victor E)
                            'Costo de Financiamiento cobrado admin'
                            ]]

#%%
#Abrimos el excel donde contiene el reporte:
if duplicados.shape[0] == 0:

    excel           = UBI + '/' + DOCUMENTO
    workbook        = load_workbook(excel)
    BD_Contabilidad = workbook['Online']
    
#%%
if duplicados.shape[0] == 0:

    for row in BD_Contabilidad.iter_rows(min_row = 2, 
                                         max_row = base_actual.shape[0]+1, 
                                         min_col = 3, 
                                         max_col = 20):  # C=3, U=20
        for cell in row:
            cell.value = None  # Borra el contenido pero mantiene el formato
    
    # Escribir el DataFrame en el Excel desde la columna C
    for i, row in enumerate(base_actual.itertuples(index = False), start = 2):  # Comenzar desde la fila 2
        for j, value in enumerate(row, start = 3):  # Escribir desde la columna C (col=3)
            BD_Contabilidad.cell(row=i, column=j, value=value)

#%% Función para eliminar un estilo si ya existe
def eliminar_estilo_si_existe(workbook, nombre_estilo):
    for estilo in workbook.named_styles:
        if isinstance(estilo, NamedStyle) and estilo.name == nombre_estilo:
            workbook.named_styles.remove(estilo)
            print(f"Estilo {nombre_estilo} eliminado.")
            break  # Salir después de eliminar el estilo

# Eliminar los estilos si existen
# eliminar_estilo_si_existe(workbook, 'date_style')
# eliminar_estilo_si_existe(workbook, 'percentage_style')

#%% Añadir nuevos registros al final
# ultima_fila = BD_Contabilidad.max_row
if duplicados.shape[0] == 0:

    ultima_fila = base_actual.shape[0]+1
    
    for i, row in nuevos_registros.iterrows():
        for col_num, value in enumerate(row, start=1):
            BD_Contabilidad.cell(row=ultima_fila + 1, column=col_num, value=value)
        ultima_fila += 1  # mover la fila después de escribir
    
#%% Crear y agregar nuevos estilos
# Crear el estilo de fecha si no existe
# date_style = NamedStyle(name="date_style", number_format='DD/MM/YYYY')
# percentage_style = NamedStyle(name="percentage_style", number_format='0.00%')

# # Aplicar el estilo de fecha a las columnas correspondientes
# columnas_fecha = ['E', 'F', 'G', 'S', 'T']
# for columna in columnas_fecha:
#     for row in BD_Contabilidad[columna]:
#         row.style = date_style

# # Aplicar el estilo de porcentaje a las columnas correspondientes
# columna_porcentaje = ['J', 'K']
# for columna in columna_porcentaje:
#     for row in BD_Contabilidad[columna]:
#         row.style = percentage_style

# # Añadir los estilos al libro
# if 'date_style' not in [estilo.name for estilo in workbook.named_styles]:
#     workbook.add_named_style(date_style)

# if 'percentage_style' not in [estilo.name for estilo in workbook.named_styles]:
#     workbook.add_named_style(percentage_style)

# Guardar los cambios
if duplicados.shape[0] == 0:

    workbook.save(excel)

print('datos insertados en Gestión de Comprobantes')
#%% columna auxiliar de garantía
copia_base['Subasta'] = copia_base['Subasta'].str.lower()
garantias['CODIGO SUBASTA'] = garantias['CODIGO SUBASTA'].str.lower()
copia_base = copia_base.merge(garantias,
                              left_on  = 'Subasta',
                              right_on = 'CODIGO SUBASTA',
                              how      = 'left')

copia_base.rename(columns = {'GARANTIA': 'GARANTIA NEGATIVA'}, inplace=True)

copia_base.rename(columns = {'int_pronto_pago': 'PRONTO PAGO ADMIN (julio 2025 en adelante)'}, inplace = True)

copia_base[['Subasta', 
            'GARANTIA NEGATIVA', 
            'PRONTO PAGO ADMIN (julio 2025 en adelante)']].to_excel(R'C:\Users\Joseph Montoya\Desktop\columna garantía y pronto pago.xlsx', 
                                           index = False )

print('creadas garantías negativas, y ops pronto pago')

#%%
# =============================================================================
# =============================================================================
# =============================================================================
# =============================================================================
# =============================================================================
# =============================================================================
# =============================================================================
# # # # # # # parte de las operaciones offline
# =============================================================================
# =============================================================================
# =============================================================================
# =============================================================================
# =============================================================================
# =============================================================================
# =============================================================================

query_offline = ''' 

with capa1 as (   
SELECT
    hd.tipo_de_producto,
    hd.tipo_de_operacion,
    HD.dealname as Codigo_de_Subasta,
    DS.label_dealstage AS Etapa_del_Negocio, 
    hd.moneda_del_monto_financiado AS Moneda_del_Monto_Financiado,
    cast(  hd.fecha_de_desembolso__factoring_ AS date) AS Fecha_Desembolso,
    CAST((hd.closedate - INTERVAL  '5' HOUR) AS DATE) AS Fecha_venta,
    hd.fecha_de_pago__factoring_ as Fecha_esperada_pago,
    hd.monto_financiado as Monto_Financiado,
    hd.monto_neto_de_facturas__factoring_ as Monto_neto,
    
    (CASE 
        WHEN CAST(REPLACE(REPLACE(hd.tasa_de_financiamiento____, '%', ''), ',', '.') AS DOUBLE) > 6E-1
            THEN round(CAST(REPLACE(REPLACE(hd.tasa_de_financiamiento____, '%', ''), ',', '.') AS DOUBLE) / 100, 5)
        ELSE round(CAST(REPLACE(REPLACE(hd.tasa_de_financiamiento____, '%', ''), ',', '.') AS DOUBLE), 5)
    END) AS Tasa_interes_empresario,
    
    round(CAST(REPLACE(REPLACE(hd.tasa_de_venta____, '%', ''), ',', '.') AS DOUBLE) / 100,5) as Tasa_interes_crowd,
    
    --'' as Costo_Financiamiento_teorico
    hd.ruc_proveedor,
    hd.proveedor as razon_social,
    ' ' as Direccion,
    ' ' as correo,
    HT.CLOSED_DATE as fecha_pago_real,
    ht.monto_pagado_facturas as  Monto_pagado_total,
    hd.comision_estructuracion
    
FROM prod_datalake_master.hubspot__deal AS HD

LEFT JOIN prod_datalake_master.ba__pipelines_id AS PID
ON PID.ID =  HD.pipeline

LEFT JOIN prod_datalake_master.ba__dealstages_id AS DS 
ON DS.id_dealstage = HD.dealstage

LEFT JOIN (select * from prod_datalake_master.hubspot__ticket WHERE hs_pipeline = '26417284') AS ht
ON HD.dealname = HT.subject

WHERE PID.LABEL = 'Prestamype - Factoring'
and ds.label_dealstage not in ( 'Cancelado (Subasta desierta) (Prestamype - Factoring)', 
                                'Rechazado (Prestamype - Factoring)',
                                'Generado por Compra (Prestamype - Factoring)',
                                'Generado por Offline Varios Fondos (Prestamype - Factoring)')
--and hd.tipo_de_operacion in ('Offline', 'Mixta', 'Ordering')
and length(hd.dealname) > 10

order by hd.dealname
), capa2 as (           
select 
    tipo_de_producto,
    Codigo_de_Subasta,
    Etapa_del_Negocio,
    tipo_de_operacion,
    Moneda_del_Monto_Financiado,
    Fecha_Desembolso,
    Fecha_venta,
    Fecha_esperada_pago,
    Fecha_Pago_real,
    Monto_neto,
    Monto_Financiado,
    Tasa_interes_empresario,
    Tasa_interes_crowd,
    
    --(date_diff('day', Fecha_Desembolso, Fecha_esperada_pago)) as dias_dif,
    
        round((Monto_Financiado * (
            pow(1 + Tasa_interes_empresario, date_diff('day', Fecha_Desembolso, Fecha_esperada_pago) / 30.0)
        ) - Monto_Financiado),2) AS Costo_Financiamiento_teorico,
        
        ruc_proveedor,
        razon_social,
        Direccion,
        correo,
        comision_estructuracion,
        
        '' AS Comprobante_Comision_manual,
        '' AS Comprobante_costo_financiamiento_manual,
        Fecha_Desembolso AS Fecha_Desembolso_Hubspot,

        '' as "Monto pagado total (manual)",
        '' as "Estado de cobranza real (manual)",
        Monto_pagado_total as "Monto pagado total (teórico para validaciones)",
        '' as "Interés Bruto pagado a Crowd (manual)",
        '' as "Costo de Financiamiento cobrado admin",
        '' AS "Costo de Financiamiento Liquidado emp(numérico)",
        '' AS "Costo de Financiamiento Liquidado emp(comentarios)",
        Moneda_del_Monto_Financiado AS "Moneda",
        ' ' AS "Monto pagado - Monto financiado"
         
        
    FROM capa1

)
select * from capa2


'''

#%%
import pandas as pd
# import numpy as np
# import boto3
from pyathena import connect
# import openpyxl
from openpyxl import load_workbook
#from openpyxl.styles import NamedStyle
import os

import shutil
from datetime import datetime

import warnings
warnings.filterwarnings("ignore")

from datetime import datetime, timezone, timedelta
peru_tz = timezone(timedelta(hours=-5))
today_date = datetime.now(peru_tz).strftime('%Y%m%d')

#%% Credenciales de AmazonAthena
import json
with open(r"C:/Users/Joseph Montoya/Desktop/credenciales actualizado.txt") as f:
    creds = json.load(f)

conn = connect(
    aws_access_key_id     = creds["AccessKeyId"],
    aws_secret_access_key = creds["SecretAccessKey"],
    aws_session_token     = creds["SessionToken"],
    s3_staging_dir        = creds["s3_staging_dir"],
    region_name           = creds["region_name"]
    
    )
#%%%
cursor = conn.cursor()
cursor.execute(query_offline)

# Obtener los resultados
resultados = cursor.fetchall()

# Obtener los nombres de las columnas
column_names = [desc[0] for desc in cursor.description]

# Convertir los resultados a un DataFrame de pandas
df_offlines = pd.DataFrame(resultados, columns=column_names)

print('query offline ejecutada')
#%% convertir columnas de tasas, donde la separación es el punto, a coma
# comprobar si este código es eliminable cuando se automatice la escritura en el google sheet

# Lista de columnas a convertir

'''
columnas = ['Tasa_interes_empresario', 
            'Tasa_interes_crowd',
            'Monto_Financiado',
            'Monto_neto',
            'Costo_Financiamiento_teorico',
            'Monto pagado total (teórico para validaciones)',
            ]

# Formatear cada columna como texto con coma decimal
for col in columnas:
    df_offlines[col] = df_offlines[col].apply(lambda x: f"{x:.5f}".replace('.', ',') if pd.notna(x) else "")
'''
#%%
df_offlines = df_offlines.fillna(' ')
#%%
df_offlines.to_excel(rf'C:\Users\Joseph Montoya\Desktop\pruebas\gestion de comprobantes offlines {today_date}.xlsx',
                     index = False)

print('excel offline creado')
#%%
print('fin')

                                                                
